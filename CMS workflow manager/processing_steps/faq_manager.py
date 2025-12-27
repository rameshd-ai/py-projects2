"""
Step 4: FAQ Manager Migration
Handles FAQ data migration from source to destination site
Similar to menu processing - fetches FAQ data and creates records in destination CMS
"""
import os
import json
import sys
import time
import logging
import requests
from typing import Dict, Any, List
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# FAQ processing only exports to Excel, does NOT create records in CMS
# No CMS API imports needed
from utils import get_job_folder, get_job_output_folder, ensure_job_folders, load_job_config
from config import BASE_DIR

# Configure UTF-8 encoding for console output on Windows
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

logger = logging.getLogger(__name__)

# Column definitions for FAQ output
COLUMN_MAPPING = {
    'Question': 'Question',  
    'Answer': 'Answer',
    'Category': 'Category',
    'ImageURL': 'Image URL',
    'ImageAltTag': 'Image Alt tag',
    'Action': 'Action',
    'QuestionId': 'QuestionId_TO_DROP', 
    'CategoryId': 'CategoryId_TO_DROP',
}

DESIRED_COLUMN_ORDER = [
    'Question', 
    'Answer', 
    'Category', 
    'Image URL', 
    'Image Alt tag', 
    'Web URL',
    'Action',
]


def load_json_data(file_path: str) -> Dict[str, Any]:
    """Load JSON data from file"""
    try:
        if os.path.exists(file_path):
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logger.error(f"Error loading JSON from {file_path}: {e}")
    return {}


def fetch_faq_data_from_source(source_link: str) -> List[Dict[str, Any]]:
    """
    Fetch FAQ data from source API endpoint
    """
    # Validate and prepare source link
    if not source_link.startswith(('http://', 'https://')):
        source_link = 'https://' + source_link
    
    parsed_url = urlparse(source_link)
    if not parsed_url.netloc:
        raise ValueError(f"Invalid source URL: {source_link}")
    
    base_url = f"{parsed_url.scheme}://{parsed_url.netloc}".rstrip('/')
    faq_api_url = f"{base_url}/api/FAQApi/GetFAQList"
    
    print(f"[FAQ] Fetching FAQ data from: {faq_api_url}", flush=True)
    logger.info(f"Fetching FAQ data from: {faq_api_url}")
    
    # Make direct API request
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36',
        'Accept': 'application/json'
    }
    
    response = requests.get(faq_api_url, headers=headers, timeout=30)
    response.raise_for_status()
    
    # Parse JSON response
    json_data = response.json()
    
    if not isinstance(json_data, list) or not json_data:
        raise Exception(f"No FAQ data found at {faq_api_url}")
    
    print(f"[OK] Fetched {len(json_data)} FAQ items from source", flush=True)
    logger.info(f"Fetched {len(json_data)} FAQ items from source")
    
    return json_data




def process_faq_from_source_link(job_id: str, source_link: str) -> Dict[str, Any]:
    """
    Process FAQ data from a source link - fetch and export to Excel
    Similar to reference implementation - only exports to Excel, does NOT create records in CMS
    """
    print("\n" + "="*80, flush=True)
    print(f"[FAQ PROCESSING] STARTING FAQ PROCESSING", flush=True)
    print("="*80, flush=True)
    logger.info(f"Processing FAQ from source link: {source_link} for job {job_id}")
    
    # Ensure job folders exist
    ensure_job_folders(job_id)
    output_dir = get_job_output_folder(job_id)
    
    try:
        # Step 1: Fetch FAQ data from source
        print(f"\n[STEP 1] Fetching FAQ data from source...", flush=True)
        faq_data = fetch_faq_data_from_source(source_link)
        print(f"[OK] Step 1 completed: Fetched {len(faq_data)} FAQ items", flush=True)
        
        # Save raw FAQ data
        job_folder = get_job_folder(job_id)
        raw_faq_file = os.path.join(job_folder, "faq_raw_data.json")
        with open(raw_faq_file, 'w', encoding='utf-8') as f:
            json.dump(faq_data, f, indent=4)
        print(f"[OK] Raw FAQ data saved to: {raw_faq_file}", flush=True)
        
        # Step 2: Create Excel export file
        print(f"\n[STEP 2] Creating Excel export file...", flush=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "FAQ Data"
        
        # Write headers
        headers_row = DESIRED_COLUMN_ORDER
        for col_idx, header in enumerate(headers_row, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Process each FAQ item
        row_num = 2
        for item in faq_data:
            row_data = {}
            for old_key, new_key in COLUMN_MAPPING.items():
                if old_key in item:
                    if new_key.endswith('_TO_DROP'):
                        continue
                    row_data[new_key] = str(item[old_key]) if item[old_key] is not None else ''
            row_data['Web URL'] = source_link
            row_data['Action'] = 'publish'
            
            for col_idx, col_name in enumerate(headers_row, start=1):
                value = row_data.get(col_name, '')
                ws.cell(row=row_num, column=col_idx, value=value)
            row_num += 1
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(headers_row, start=1):
            max_length = len(col_name)
            for row in ws.iter_rows(min_row=2, max_row=row_num-1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
        
        output_filename = f"faq_processed_{job_id}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        wb.save(output_path)
        print(f"[OK] Step 2 completed: Excel file saved to {output_filename}", flush=True)
        
        processed_count = len(faq_data)
        print(f"\n[SUCCESS] FAQ processing completed: {processed_count} FAQ items exported to Excel", flush=True)
        logger.info(f"FAQ processing completed. Processed {processed_count} FAQ items")
        
        return {
            "success": True,
            "processed_count": processed_count,
            "output_file": output_filename,
            "file_path": output_path,
            "message": f"Successfully processed {processed_count} FAQ items and exported to Excel"
        }
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Failed to fetch FAQ data: {e}")
        raise Exception(f"Failed to fetch FAQ data from API: {e}")
    except Exception as e:
        logger.error(f"FAQ processing failed: {e}")
        raise Exception(f"FAQ processing failed: {e}")


def run_faq_manager_step(job_id: str, step_config: Dict, workflow_context: Dict) -> Dict[str, Any]:
    """
    Step 4: FAQ Manager Migration
    Migrates FAQ data from source to destination site
    """
    logger.info(f"Starting FAQ Manager migration for job {job_id}")
    
    job_config = workflow_context.get("job_config", {})
    
    # Check if FAQ Manager is enabled
    if not job_config.get("faqManager", False):
        return {
            "faq_migration_enabled": False,
            "message": "FAQ Manager migration skipped (not enabled)"
        }
    
    # Get source link from workflow context or job config
    source_link = workflow_context.get("source_link") or job_config.get("faq_source_link")
    
    if not source_link:
        raise ValueError("Source link is required for FAQ processing")
    
    try:
        # Process FAQ from source link
        result = process_faq_from_source_link(job_id, source_link)
        
        return {
            "faq_migration_enabled": True,
            "faq_migrated": True,
            **result
        }
        
    except Exception as e:
        logger.error(f"FAQ Manager migration failed: {e}")
        raise Exception(f"FAQ Manager migration failed: {e}")
